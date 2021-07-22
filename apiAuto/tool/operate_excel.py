#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time : 2021/7/14 16:10
# @Author : lumi
# @File : operate_excel.py
# @Software: PyCharm
import openpyxl

class OpExcel:

    def __init__(self,file_path,sheetname):
        self.book = openpyxl.load_workbook(file_path)
        self.work_sheet = self.book[sheetname]

    def read_data(self,row,column):

        return self.work_sheet.cell(row=row, column=column).value

    def get_max_row(self):
        return self.work_sheet.max_row



